import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment


# ============================================================
# Handle Arguments
# ============================================================

if len(sys.argv) != 3:
    print("Usage: generate_excel.py <benchstat.txt> <output.xlsx>")
    sys.exit(1)

INPUT_FILE = Path(sys.argv[1])
OUTPUT_FILE = Path(sys.argv[2])

lines = INPUT_FILE.read_text().splitlines()


# ============================================================
# Helper Functions
# ============================================================

def format_percent(val):
    if val is None:
        return "%nan"
    return f"{val:+.2f}%"


def extract_geomean(section):

    for line in section:

        line = line.strip()

        if line.startswith("geomean"):

            m = re.search(r'([+-]?\d+\.\d+)%', line)

            if m:
                return float(m.group(1))

    return None


def extract_metric_map(section):

    data = {}

    for line in section:

        line = line.strip()

        if not line:
            continue

        if line.startswith("geomean"):
            continue

        if line.startswith("│"):
            continue

        if line.startswith("¹"):
            continue

        if line.startswith("²"):
            continue

        if line.startswith("³"):
            continue

        bench_match = re.match(r'^(\S+)', line)

        if not bench_match:
            continue

        benchmark = bench_match.group(1)

        pct_match = re.search(r'([+-]\d+\.\d+)%', line)

        if pct_match:
            data[benchmark] = float(pct_match.group(1))
        else:
            data[benchmark] = None

    return data


def find_metric_sections(block):

    sections = {
        "sec/op": [],
        "B/op": [],
        "allocs/op": []
    }

    current = None

    for line in block:

        if "│" in line and "sec/op" in line:
            current = "sec/op"
            continue

        if "│" in line and "B/op" in line:
            current = "B/op"
            continue

        if "│" in line and "allocs/op" in line:
            current = "allocs/op"
            continue

        if current:
            sections[current].append(line)

    return sections


# ============================================================
# Parse package blocks
# ============================================================

blocks = []
current_block = []

for line in lines:

    if line.startswith("pkg:"):

        if current_block:
            blocks.append(current_block)

        current_block = [line]

    else:
        current_block.append(line)

if current_block:
    blocks.append(current_block)


package_summary = []
package_details = []

for block in blocks:

    pkg_name = block[0].replace("pkg:", "").strip()

    sections = find_metric_sections(block)

    sec_section = sections["sec/op"]
    b_section = sections["B/op"]
    alloc_section = sections["allocs/op"]

    sec_geo = extract_geomean(sec_section)
    b_geo = extract_geomean(b_section)
    alloc_geo = extract_geomean(alloc_section)

    package_summary.append({
        "Package": pkg_name,
        "sec/op": sec_geo,
        "B/op": b_geo,
        "allocs/op": alloc_geo
    })

    sec_map = extract_metric_map(sec_section)
    b_map = extract_metric_map(b_section)
    alloc_map = extract_metric_map(alloc_section)

    benchmark_names = (
        set(sec_map.keys()) |
        set(b_map.keys()) |
        set(alloc_map.keys())
    )

    benchmark_rows = []

    for benchmark in sorted(benchmark_names):

        benchmark_rows.append({
            "Benchmark": benchmark,
            "sec/op": sec_map.get(benchmark),
            "B/op": b_map.get(benchmark),
            "allocs/op": alloc_map.get(benchmark)
        })

    package_details.append({
        "Package": pkg_name,
        "sec/op": sec_geo,
        "B/op": b_geo,
        "allocs/op": alloc_geo,
        "Benchmarks": benchmark_rows
    })


# ============================================================
# Write Excel
# ============================================================

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:

    # --------------------------------------------------------
    # Package Summary
    # --------------------------------------------------------

    summary_df = pd.DataFrame(package_summary)

    summary_df.to_excel(
        writer,
        sheet_name="Package Summary",
        index=False
    )

    workbook = writer.book

    header_fill = PatternFill(
        fill_type="solid",
        start_color="5B9BD5",
        end_color="5B9BD5"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
        size=13
    )

    package_fill = PatternFill(
        fill_type="solid",
        start_color="A9D18E",
        end_color="A9D18E"
    )

    package_font = Font(
        bold=True,
        color="000000",
        size=14
    )

    center_align = Alignment(horizontal="center")

    # --------------------------------------------------------
    # Format Summary Sheet
    # --------------------------------------------------------

    summary_ws = workbook["Package Summary"]

    for cell in summary_ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # --------------------------------------------------------
    # Benchmark Details Sheet
    # --------------------------------------------------------

    ws = workbook.create_sheet("Benchmark Details")

    row = 1

    for pkg in package_details:

        header_cell = ws.cell(row=row, column=1)
        header_cell.value = f"PACKAGE: {pkg['Package']}"
        header_cell.font = package_font
        header_cell.fill = package_fill

        row += 1

        ws.cell(
            row=row,
            column=1
        ).value = (
            f"Geomean: "
            f"sec/op={format_percent(pkg['sec/op'])}   "
            f"B/op={format_percent(pkg['B/op'])}   "
            f"allocs/op={format_percent(pkg['allocs/op'])}"
        )

        row += 2

        headers = [
            "Benchmark",
            "sec/op",
            "B/op",
            "allocs/op"
        ]

        for col, header in enumerate(headers, start=1):

            cell = ws.cell(row=row, column=col)

            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row += 1

        for bench in pkg["Benchmarks"]:

            ws.cell(row=row, column=1).value = bench["Benchmark"]
            ws.cell(row=row, column=2).value = format_percent(bench["sec/op"])
            ws.cell(row=row, column=3).value = format_percent(bench["B/op"])
            ws.cell(row=row, column=4).value = format_percent(bench["allocs/op"])

            row += 1

        row += 3

    # --------------------------------------------------------
    # Auto-size columns
    # --------------------------------------------------------

    for sheet in workbook.worksheets:

        for col in sheet.columns:

            max_len = 0

            for cell in col:

                try:
                    max_len = max(
                        max_len,
                        len(str(cell.value))
                    )
                except Exception:
                    pass

            sheet.column_dimensions[
                col[0].column_letter
            ].width = max_len + 5

print(f"Excel report generated: {OUTPUT_FILE}")
